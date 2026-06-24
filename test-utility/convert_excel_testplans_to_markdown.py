from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "resources" / "test-cases"
OUTPUT_DIR = ROOT / "test-cases-parsed"


def normalize_output_name(workbook_path: Path) -> str:
    base = workbook_path.stem
    base = re.sub(r"\s*-\s*Automated\s+Test\s+Plan\s*$", "", base, flags=re.IGNORECASE)
    base = re.sub(r"\s*-\s*Test\s+Plan\s*$", "", base, flags=re.IGNORECASE)
    base = re.sub(r"\s+Test\s+Plan\s*$", "", base, flags=re.IGNORECASE)
    base = re.sub(r"\s+", "-", base.strip())
    base = re.sub(r"-+", "-", base)
    return f"{base}-TestPlan.md"


def detect_section(headers: list[str], worksheet_name: str) -> str:
    text = " ".join(headers + [worksheet_name]).lower()
    if "execution" in text or "result" in text or "tester" in text:
        return "Test Execution"
    if "requirement" in text or "work item" in text:
        return "Requirements"
    if "test case" in text or "expected result" in text or "precondition" in text:
        return "Test Cases"
    return "Data"


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def escape_markdown(text: str) -> str:
    text = text.replace("\\", "\\\\")
    text = text.replace("|", "\\|")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\n", "<br>")
    return text


def is_blank(text: str) -> bool:
    return text.strip() == ""


def trim_empty_edges(matrix: list[list[str]]) -> list[list[str]]:
    if not matrix:
        return matrix

    non_empty_rows = [i for i, row in enumerate(matrix) if any(not is_blank(v) for v in row)]
    if not non_empty_rows:
        return []

    non_empty_cols = [
        j for j in range(len(matrix[0])) if any(not is_blank(matrix[i][j]) for i in range(len(matrix)))
    ]
    if not non_empty_cols:
        return []

    return [[row[j] for j in non_empty_cols] for i, row in enumerate(matrix) if i in non_empty_rows]


def matrix_from_sheet(ws) -> list[list[str]]:
    non_empty = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                non_empty.append((cell.row, cell.column))

    if not non_empty:
        return []

    min_row = min(r for r, _ in non_empty)
    max_row = max(r for r, _ in non_empty)
    min_col = min(c for _, c in non_empty)
    max_col = max(c for _, c in non_empty)

    matrix: list[list[str]] = []
    for r in range(min_row, max_row + 1):
        row_values: list[str] = []
        for c in range(min_col, max_col + 1):
            row_values.append(cell_to_text(ws.cell(row=r, column=c).value))
        matrix.append(row_values)

    return trim_empty_edges(matrix)


def markdown_table(matrix: list[list[str]]) -> str:
    if not matrix:
        return ""

    header = [h.strip() if h.strip() else " " for h in matrix[0]]
    data_rows = matrix[1:] if len(matrix) > 1 else []

    lines = []
    lines.append("| " + " | ".join(escape_markdown(h) for h in header) + " |")
    lines.append("|" + "|".join("---" for _ in header) + "|")

    if data_rows:
        for row in data_rows:
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(escape_markdown(v) for v in padded[: len(header)]) + " |")
    else:
        lines.append("| " + " | ".join("" for _ in header) + " |")

    return "\n".join(lines)


def row_non_empty_cells(row: list[str]) -> list[str]:
    return [cell for cell in row if not is_blank(cell)]


def row_header_score(row: list[str]) -> int:
    keywords = {
        "id",
        "title",
        "priority",
        "precondition",
        "step",
        "expected",
        "result",
        "requirement",
        "description",
        "owner",
        "status",
        "date",
        "tester",
        "comment",
        "risk",
        "note",
        "work item",
        "state",
        "tag",
    }
    score = 0
    for cell in row_non_empty_cells(row):
        text = cell.strip().lower()
        if any(key in text for key in keywords):
            score += 1
    return score


def detect_header_index(matrix: list[list[str]]) -> int | None:
    for idx, row in enumerate(matrix):
        non_empty = row_non_empty_cells(row)
        if not non_empty:
            continue

        score = row_header_score(row)
        if score >= 2 and len(non_empty) >= 3:
            return idx
        if len(non_empty) >= 5:
            return idx

    return None


def summarize_preamble_rows(rows: list[list[str]]) -> list[str]:
    summary: list[str] = []
    for row in rows:
        values = [escape_markdown(cell.strip()) for cell in row if not is_blank(cell)]
        if values:
            summary.append(" | ".join(values))
    return summary


def gather_notes(ws) -> list[str]:
    notes: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.comment and cell.comment.text:
                note = cell.comment.text.strip()
                if note:
                    notes.append(note)
    return notes


def worksheet_to_markdown(ws) -> str:
    matrix = matrix_from_sheet(ws)
    lines: list[str] = [f"## {ws.title}", "", "### Overview", ""]

    if not matrix:
        lines.append("> This worksheet is empty.")
        lines.append("")
    else:
        header_idx = detect_header_index(matrix)
        preamble_rows: list[list[str]] = []
        table_matrix = matrix

        if header_idx is not None and header_idx > 0:
            preamble_rows = matrix[:header_idx]
            table_matrix = matrix[header_idx:]

        if preamble_rows:
            lines.append("### Metadata")
            lines.append("")
            for item in summarize_preamble_rows(preamble_rows):
                lines.append(f"- {item}")
            lines.append("")

        headers = [cell_to_text(v) for v in table_matrix[0]] if table_matrix else []
        section = detect_section(headers, ws.title)
        lines.append(f"### {section}")
        lines.append("")
        lines.append(markdown_table(table_matrix))
        lines.append("")

    notes = gather_notes(ws)
    if notes:
        lines.append("### Notes")
        lines.append("")
        for note in notes:
            lines.append(f"- {escape_markdown(note)}")
        lines.append("")

    merged = [str(rng) for rng in ws.merged_cells.ranges]
    if merged:
        lines.append("### Merged Cells")
        lines.append("")
        for rng in merged:
            lines.append(f"- {rng}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def workbook_to_markdown(workbook_path: Path) -> str:
    wb = load_workbook(workbook_path, data_only=False)
    lines: list[str] = [f"# {workbook_path.stem}", ""]
    for ws in wb.worksheets:
        lines.append(worksheet_to_markdown(ws))
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def iter_workbooks(folder: Path) -> Iterable[Path]:
    return sorted(folder.glob("*.xlsx"), key=lambda p: p.name.lower())


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbooks = list(iter_workbooks(INPUT_DIR))
    if not workbooks:
        print(f"No .xlsx files found in: {INPUT_DIR}")
        return

    for workbook_path in workbooks:
        md = workbook_to_markdown(workbook_path)
        output_name = normalize_output_name(workbook_path)
        output_path = OUTPUT_DIR / output_name
        output_path.write_text(md, encoding="utf-8")
        print(f"Converted: {workbook_path.name} -> {output_path.name}")

    print(f"Done. Converted {len(workbooks)} workbook(s).")


if __name__ == "__main__":
    main()
